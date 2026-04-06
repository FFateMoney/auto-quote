from __future__ import annotations

import base64
import io
import logging
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenPyxlImage
from openpyxl.utils import get_column_letter
from PIL import Image

from packages.core.logging_utils import append_run_log
from packages.core.models import DocumentAsset, NormalizedTextBlock

from .ocr_adapter import OcrAdapter


logger = logging.getLogger(__name__)
MAX_IMAGE_BYTES = 10 * 1024 * 1024
SUPPORTED_IMAGE_TYPES = {
    "png": "image/png",
    "jpg": "image/jpeg",
    "jpeg": "image/jpeg",
    "bmp": "image/bmp",
    "gif": "image/gif",
    "webp": "image/webp",
    "tiff": "image/tiff",
}
IMAGE_PLACEHOLDER_RE = re.compile(r"\[(IMAGE_\d+)\]")


@dataclass(slots=True)
class ExcelNormalizedPayload:
    text_blocks: list[NormalizedTextBlock]
    assets: list[DocumentAsset]
    metadata: dict[str, Any]


class ExcelAdapter:
    def __init__(self, *, ocr_adapter: OcrAdapter | None = None) -> None:
        self.ocr_adapter = ocr_adapter or OcrAdapter()

    def extract(self, input_path: Path, *, run_dir: Path) -> ExcelNormalizedPayload:
        logger.info("Excel 预处理开始: file=%s", input_path.name)
        append_run_log(run_dir, f"Excel 预处理开始: {input_path.name}")

        workbook = load_workbook(filename=input_path, data_only=True)
        extracted_dir = run_dir / "excel_images" / input_path.stem
        extracted_dir.mkdir(parents=True, exist_ok=True)

        text_blocks: list[NormalizedTextBlock] = []
        assets: list[DocumentAsset] = []
        metadata: dict[str, Any] = {
            "sheet_count": len(workbook.worksheets),
            "row_count": 0,
            "image_count": 0,
        }

        for sheet in workbook.worksheets:
            sheet_blocks, sheet_assets, sheet_row_count = self._extract_sheet(
                sheet=sheet,
                extracted_dir=extracted_dir,
                asset_offset=len(assets),
                run_dir=run_dir,
            )
            text_blocks.extend(sheet_blocks)
            assets.extend(sheet_assets)
            metadata["row_count"] += sheet_row_count
            metadata["image_count"] += len(sheet_assets)

        if metadata["image_count"] > 250:
            raise RuntimeError("excel_image_count_exceeded")

        append_run_log(
            run_dir,
            f"Excel 预处理完成: {input_path.name} | rows={metadata['row_count']} | images={metadata['image_count']}",
        )
        return ExcelNormalizedPayload(text_blocks=text_blocks, assets=assets, metadata=metadata)

    def _extract_sheet(
        self,
        *,
        sheet: Any,
        extracted_dir: Path,
        asset_offset: int,
        run_dir: Path,
    ) -> tuple[list[NormalizedTextBlock], list[DocumentAsset], int]:
        row_images = self._extract_sheet_images(sheet=sheet, extracted_dir=extracted_dir, asset_offset=asset_offset, run_dir=run_dir)
        row_values: dict[int, list[str]] = {}
        header_candidate_rows: set[int] = set()
        header_row_index: int | None = None
        headers: list[str] = []
        row_count = 0

        for row in sheet.iter_rows():
            row_index = row[0].row if row else 0
            values = [self._cell_text(cell.value) for cell in row]
            if not any(values) and row_index not in row_images:
                continue

            if header_row_index is None and any(values):
                header_row_index = row_index
                headers = self._build_headers(values)
                row_values[row_index] = self._rendered_pairs(values)
                header_candidate_rows.add(row_index)
                continue

            row_values[row_index] = self._rendered_pairs(values, headers=headers)

        blocks: list[NormalizedTextBlock] = []
        all_row_indexes = sorted(set(row_values) | set(row_images))
        for row_index in all_row_indexes:
            rendered_pairs = row_values.get(row_index, [])
            image_items = row_images.get(row_index, [])
            fragments = [
                f"[SHEET] {sheet.title}",
                f"[ROW] {row_index}",
            ]
            if row_index in header_candidate_rows:
                fragments.append("[HEADER_CANDIDATE] true")
            fragments.extend(rendered_pairs)
            if image_items:
                for item in image_items:
                    if item["ocr_text"]:
                        fragments.append(f"图片OCR({item['asset'].asset_id})={item['ocr_text']}")
                    fragments.append(f"图片引用=[{item['asset'].asset_id}]")
            block_text = "\n".join(fragment for fragment in fragments if fragment).strip()
            if not block_text:
                continue
            context_text = self._context_text(block_text)
            for item in image_items:
                item["asset"].context_text = context_text
            blocks.append(
                NormalizedTextBlock(
                    block_id=f"{sheet.title}-row-{row_index}",
                    block_type="ExcelRow",
                    text=block_text,
                    source_path=f"{sheet.title}!{row_index}",
                )
            )
            row_count += 1

        assets = [item["asset"] for row_index in all_row_indexes for item in row_images.get(row_index, [])]
        return blocks, assets, row_count

    def _rendered_pairs(self, values: list[str], *, headers: list[str] | None = None) -> list[str]:
        rendered_pairs: list[str] = []
        current_headers = headers or [f"{get_column_letter(idx + 1)}列" for idx in range(len(values))]
        for idx, value in enumerate(values):
            if not value:
                continue
            label = current_headers[idx] if idx < len(current_headers) else f"{get_column_letter(idx + 1)}列"
            rendered_pairs.append(f"{label}={value}")
        return rendered_pairs

    def _extract_sheet_images(
        self,
        *,
        sheet: Any,
        extracted_dir: Path,
        asset_offset: int,
        run_dir: Path,
    ) -> dict[int, list[dict[str, Any]]]:
        by_row: dict[int, list[dict[str, Any]]] = defaultdict(list)
        images = list(getattr(sheet, "_images", []) or [])
        for index, image in enumerate(images, start=1):
            row_index, col_index = self._anchor_position(image)
            raw_bytes = self._image_bytes(image)
            asset = self._build_asset(
                image=image,
                raw_bytes=raw_bytes,
                seq=asset_offset + sum(len(items) for items in by_row.values()) + 1,
                sheet_title=sheet.title,
                row_index=row_index,
                col_index=col_index,
            )
            temp_path = extracted_dir / f"{asset.asset_id}.{self._image_extension(image)}"
            temp_path.write_bytes(raw_bytes)
            ocr = self.ocr_adapter.extract(temp_path, run_dir=run_dir)
            ocr_text = self._ocr_text(ocr.full_text)
            by_row[row_index].append({"asset": asset, "ocr_text": ocr_text, "image_index": index})
        return by_row

    def _build_asset(
        self,
        *,
        image: OpenPyxlImage,
        raw_bytes: bytes,
        seq: int,
        sheet_title: str,
        row_index: int,
        col_index: int,
    ) -> DocumentAsset:
        if len(raw_bytes) > MAX_IMAGE_BYTES:
            raise RuntimeError(f"excel_image_size_exceeded:{sheet_title}:{row_index}:{col_index}")
        mime_type = self._mime_type(image)
        with Image.open(io.BytesIO(raw_bytes)) as pil_image:
            width, height = pil_image.size
        if width <= 10 or height <= 10:
            raise RuntimeError(f"excel_image_too_small:{sheet_title}:{row_index}:{col_index}")
        data = base64.b64encode(raw_bytes).decode("utf-8")
        return DocumentAsset(
            asset_id=f"IMAGE_{seq}",
            mime_type=mime_type,
            data_url=f"data:{mime_type};base64,{data}",
            position=f"sheet={sheet_title},row={row_index},col={get_column_letter(col_index)}",
            context_text="",
        )

    def _image_bytes(self, image: OpenPyxlImage) -> bytes:
        payload = image._data()
        if isinstance(payload, memoryview):
            return payload.tobytes()
        return bytes(payload)

    def _image_extension(self, image: OpenPyxlImage) -> str:
        ext = str(getattr(image, "format", "") or "").lower().strip(".")
        return ext if ext in SUPPORTED_IMAGE_TYPES else "png"

    def _mime_type(self, image: OpenPyxlImage) -> str:
        return SUPPORTED_IMAGE_TYPES.get(self._image_extension(image), "image/png")

    def _anchor_position(self, image: OpenPyxlImage) -> tuple[int, int]:
        anchor = getattr(image, "anchor", None)
        marker = getattr(anchor, "_from", None)
        if marker is None:
            return 1, 1
        return int(getattr(marker, "row", 0)) + 1, int(getattr(marker, "col", 0)) + 1

    def _build_headers(self, values: list[str]) -> list[str]:
        headers: list[str] = []
        used: set[str] = set()
        for index, value in enumerate(values):
            base = value or f"{get_column_letter(index + 1)}列"
            candidate = base
            suffix = 2
            while candidate in used:
                candidate = f"{base}_{suffix}"
                suffix += 1
            used.add(candidate)
            headers.append(candidate)
        return headers

    def _cell_text(self, value: Any) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        return re.sub(r"\s+", " ", text)

    def _ocr_text(self, value: str) -> str:
        return re.sub(r"\s+", " ", str(value or "")).strip()

    def _context_text(self, value: str) -> str:
        lines: list[str] = []
        for line in str(value or "").splitlines():
            clean = IMAGE_PLACEHOLDER_RE.sub("", line).strip()
            if not clean:
                continue
            if clean.endswith("="):
                continue
            lines.append(clean)
        return re.sub(r"\s+", " ", "\n".join(lines)).strip()
